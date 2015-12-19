# -*- coding: utf8-*-
 
from openpyxl import Workbook
from openpyxl import load_workbook
import json
wb = Workbook()
lw = load_workbook('reportar.xlsx')
dados = []
 
for i in lw.active:
    dados.append(i[0].value)
dados = set(dados)
updates = ""
tamanho = len(dados)
cont = 1
erros = []
for i in dados:
     
    print
    print "%d/%d - Porcentagem do Trabalho Finalizado: %.2f %%"%(cont,tamanho, (cont*100)/float(tamanho))
    print
    print i
    print
    novo = i.split('],[')
    novo[0] +="]]"
    print "Sugerido:"
    print novo[0]
    print
     
    if(raw_input("Deseja substituir pelo sugerido(1 - Sim / 2 - Nao)?") == "1"):
        updates += "UPDATE ACT_HI_VARINST json_dados_autor SET json_dados_autor.TEXT_ = '"+ novo[0]+"' WHERE json_dados_autor.NAME_ = 'json_dados_autor' "+"AND json_dados_autor.TEXT_ = "+ "'"+i+"';\n"
        updates += "UPDATE ACT_RU_VARIABLE json_dados_autor SET json_dados_autor.TEXT_ = '"+ novo[0]+"' WHERE json_dados_autor.NAME_ = 'json_dados_autor' "+"AND json_dados_autor.TEXT_ = "+ "'"+i+"';\n"
        updates += "UPDATE ACT_HI_DETAIL json_dados_autor SET json_dados_autor.TEXT_ = '"+ novo[0]+"' WHERE json_dados_autor.NAME_ = 'json_dados_autor' "+"AND json_dados_autor.TEXT_ = "+ "'"+i+"';"
    else:
        valor = raw_input("Digite o valor que quer substituir:")
        updates += "UPDATE ACT_HI_VARINST json_dados_autor SET json_dados_autor.TEXT_ = '"+novo[0]+"' WHERE json_dados_autor.NAME_ = 'json_dados_autor' "+"AND json_dados_autor.TEXT_ = "+ "'"+i+"';\n"
        updates += "UPDATE ACT_RU_VARIABLE json_dados_autor SET json_dados_autor.TEXT_ = '"+ novo[0]+"' WHERE json_dados_autor.NAME_ = 'json_dados_autor' "+"AND json_dados_autor.TEXT_ = "+ "'"+i+"';\n"
        updates += "UPDATE ACT_HI_DETAIL json_dados_autor SET json_dados_autor.TEXT_ = '"+ novo[0]+"' WHERE json_dados_autor.NAME_ = 'json_dados_autor' "+"AND json_dados_autor.TEXT_ = "+ "'"+i+"';"
    updates +="\n"
    cont +=1
    file = open('updates.sql','a+')
    try:
        print "Falha ao gravar no arquivo!"
        file.write(updates)
    except:
        erros.append(updates)
    file.close()
    updates=""
for  i in erros:
    print i
