# -*- coding: cp1252 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import json
wb = Workbook()
w_b = Workbook()
update_possivel = wb.active
reportar = w_b.active
lw = load_workbook('JEC - Autores Duplicados em Processos - Definição do turbilet - Turbina.xlsx')
update_possivel.append(["Numero da Pasta","Numero do Processo","Dados do Autor","PROC_NAME","TASK_NAME","ID da Tarefa","Chave do Processo","ID do Processo"])
reportar.append(["Numero da Pasta","Numero do Processo","Dados do Autor","PROC_NAME","TASK_NAME","ID da Tarefa","Chave do Processo","ID do Processo"])
count = 0

for i in lw.active:
    
    count +=1
    d = json.loads(i[2].value)
     
    teste = 0
    for j in d:
        if teste == 1:
            break
        nome = j[1]
        cpf = j[2]
        for k in d:
             
            if d.index(k) <= d.index(j):
                continue
            if(nome.strip() != k[1].strip() or cpf.strip() != k[2].strip()):
                print "Nome:",nome.strip(),"CPF:",cpf.strip()
                print "Nome:",k[1].strip(),"CPF:",k[2].strip()
                teste = 1
                break
         
         
    if(teste == 0):
        update_possivel.append(i)
        wb.save("update_possivel.xlsx")
    else:
        valor = int(raw_input("Registro:%d - Digite 1 para update e 2 para reportar:"%count))
        if(valor == 1):
            update_possivel.append(i)
            wb.save("update_possivel.xlsx")
        else:
            reportar.append(i)
            w_b.save("reportar.xlsx")
         
wb.save("update_possivel.xlsx")
w_b.save("reportar.xlsx")
